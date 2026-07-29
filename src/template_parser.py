# -*- coding: utf-8 -*-
"""
文档模板解析器
支持用户上传 .md / .txt / .docx / .xlsx 格式的自定义文档模板，
提取纯文本内容供 Prompt 模板引用。
"""

import logging
import traceback
from typing import Optional

_logger = logging.getLogger("template_parser")


def parse_template(uploaded_file) -> Optional[str]:
    """
    解析用户上传的模板文件，返回纯文本内容。

    支持格式：
    - .md / .txt / .text / .rst → 直接读取文本
    - .docx → 使用 python-docx 提取段落文本和表格
    - .xlsx → 使用 openpyxl 提取所有 Sheet 的表格内容

    Args:
        uploaded_file: 类文件对象，需支持 .name 属性和 .read() 方法（BytesIO / UploadFile）

    Returns:
        模板文本内容；解析失败返回 None
    """
    if uploaded_file is None:
        return None

    filename = uploaded_file.name.lower()

    # ---- 纯文本类格式 ----
    if filename.endswith((".md", ".txt", ".text", ".rst")):
        raw = uploaded_file.read()
        try:
            return raw.decode("utf-8")
        except UnicodeDecodeError:
            try:
                return raw.decode("gbk")
            except UnicodeDecodeError:
                return None

    # ---- Word / Excel 格式（传入类文件对象，避免二次内存拷贝）----
    if filename.endswith(".docx"):
        return _parse_docx(uploaded_file)

    if filename.endswith(".xlsx"):
        return _parse_excel(uploaded_file)

    return None


def _heading_prefix(para) -> str:
    """根据段落样式返回 Markdown 标题前缀（非标题返回空串）。

    兼容英文内置样式名（Title / Heading 1..9）与中文本地化样式名（标题 1..9）。
    """
    try:
        name = (para.style.name or "").strip()
    except Exception:
        return ""

    lower = name.lower()
    if lower == "title":
        return "# "

    level = None
    if lower.startswith("heading "):
        suffix = lower[len("heading "):].strip()
        if suffix.isdigit():
            level = int(suffix)
    elif name.startswith("标题"):
        suffix = name[len("标题"):].strip()
        if suffix.isdigit():
            level = int(suffix)

    if level and 1 <= level <= 9:
        return "#" * min(level, 6) + " "
    return ""


def _docx_table_to_markdown(table) -> list:
    """将 docx 表格转为 Markdown 行列表；空表返回空列表。"""
    rows = []
    for row in table.rows:
        cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
        rows.append(cells)

    if not rows:
        return []

    header = rows[0]
    if not header or all(not c for c in header):
        return []

    lines = ["| " + " | ".join(header) + " |",
             "| " + " | ".join(["---"] * len(header)) + " |"]
    for row_data in rows[1:]:
        # 补齐列数（处理合并单元格导致的列数不一致）
        while len(row_data) < len(header):
            row_data.append("")
        lines.append("| " + " | ".join(row_data[: len(header)]) + " |")
    lines.append("")  # 表格后空行
    return lines


def _parse_docx(file_obj) -> Optional[str]:
    """从 .docx 文件对象（BytesIO / 文件路径）中提取段落和表格文本。

    按文档 body 的原始顺序遍历段落与表格（表格保留在所属章节位置），
    并将 Heading/标题 样式转换为 Markdown 标题层级。
    """
    try:
        from docx import Document
        from docx.oxml.ns import qn
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError:
        _logger.error("python-docx 未安装，无法解析 .docx 模板")
        return None

    try:
        doc = Document(file_obj)
        parts = []

        # 按 body 子元素顺序遍历：w:p → 段落，w:tbl → 表格
        for child in doc.element.body.iterchildren():
            if child.tag == qn("w:p"):
                para = Paragraph(child, doc)
                text = para.text.strip()
                if text:
                    parts.append(_heading_prefix(para) + text)
            elif child.tag == qn("w:tbl"):
                try:
                    parts.extend(_docx_table_to_markdown(Table(child, doc)))
                except Exception:
                    _logger.warning("解析表格时出错，已跳过:\n%s", traceback.format_exc())

        result = "\n".join(parts) if parts else None
        if result:
            _logger.info("docx 解析成功: %d 段落/表格项, %d 字符", len(parts), len(result))
        return result

    except Exception:
        _logger.error("docx 解析失败:\n%s", traceback.format_exc())
        return None


def get_supported_extensions() -> list:
    """返回支持的模板文件扩展名列表。"""
    return ["md", "txt", "text", "rst", "docx", "xlsx"]


def _promote_excel_title(rows: list) -> tuple:
    """若首行为合并单元格大标题（仅 1 个非空单元格）且次行是多列表头，则将首行升格为标题。

    Returns:
        (title, rows)：title 为标题文本（无则为空串），rows 为剩余行。
    """
    if len(rows) >= 2:
        first_non_empty = [c for c in rows[0] if c]
        second_non_empty = [c for c in rows[1] if c]
        if len(first_non_empty) == 1 and len(second_non_empty) >= 2:
            return first_non_empty[0], rows[1:]
    return "", rows


def _is_enum_dict_sheet(data_rows: list) -> bool:
    """判定是否为「枚举字典」Sheet：各列是互相独立、自顶向下排列的取值清单。

    特征：每列非空单元格从首个数据行起连续排列（顶部对齐）、各列长度参差不齐、
    数据区空单元格占比高。此类 Sheet 转表格会产生行级伪关联，应转为清单式。
    """
    if len(data_rows) < 2:
        return False

    n_cols = len(data_rows[0])
    col_counts = []
    total_cells = 0
    empty_cells = 0
    for col in range(n_cols):
        values = [row[col] for row in data_rows]
        total_cells += len(values)
        empty_cells += sum(1 for v in values if not v)
        count = sum(1 for v in values if v)
        # 顶部对齐检查：非空单元格必须连续出现在列首（有间隙则视为普通记录表）
        if any(v for v in values[count:]):
            return False
        col_counts.append(count)

    filled = [c for c in col_counts if c]
    if len(filled) < 2 or max(filled) == min(filled):
        return False
    return empty_cells / total_cells >= 0.3


def _enum_sheet_lines(header: list, data_rows: list) -> list:
    """将枚举字典 Sheet 转为「字段: 值1 / 值2 ...」的清单行。"""
    lines = []
    for col, name in enumerate(header):
        values = [row[col] for row in data_rows if row[col]]
        if name and values:
            lines.append(f"- {name}: " + " / ".join(values))
    return lines


def _parse_excel(file_obj) -> Optional[str]:
    """
    从 .xlsx 文件对象（BytesIO / 文件路径）中提取所有 Sheet 的表格内容。
    每个 Sheet 以 Sheet 名作为标题；合并单元格大标题升格为独立标题行；
    枚举字典类 Sheet 转为清单式，其余转为 Markdown 表格。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        _logger.error("openpyxl 未安装，无法解析 .xlsx 模板")
        return None

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []

            for row in ws.iter_rows(values_only=True):
                # 将每个单元格转为字符串，None 转为空串
                cells = [str(cell).strip().replace("\n", " ") if cell is not None else "" for cell in row]
                # 跳过全空行
                if any(cells):
                    rows.append(cells)

            if not rows:
                continue

            # Sheet 标题
            parts.append(f"## Sheet: {sheet_name}\n")

            # 统一列数（取最大列数）
            max_cols = max(len(r) for r in rows)
            for r in rows:
                while len(r) < max_cols:
                    r.append("")

            # 首行若为合并单元格大标题，升格为独立标题行，避免占用表头
            title, rows = _promote_excel_title(rows)
            if title:
                parts.append(f"### {title}\n")

            header = rows[0]
            data_rows = rows[1:]

            if _is_enum_dict_sheet(data_rows):
                # 枚举字典 Sheet：各列独立取值清单，转表格会产生行级伪关联
                parts.extend(_enum_sheet_lines(header, data_rows))
            else:
                parts.append("| " + " | ".join(header) + " |")
                parts.append("| " + " | ".join(["---"] * max_cols) + " |")
                for row_data in data_rows:
                    parts.append("| " + " | ".join(row_data) + " |")

            parts.append("")  # Sheet 之间空行

        wb.close()
        result = "\n".join(parts) if parts else None
        if result:
            _logger.info("xlsx 解析成功: %d sheets, %d 字符", len(wb.sheetnames), len(result))
        return result

    except Exception:
        _logger.error("xlsx 解析失败:\n%s", traceback.format_exc())
        return None
